import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = load_workbook('/mnt/agents/output/FitIntel_PRO_TestCases_v3.4.xlsx')
ws = wb['Test Cases v3.4']

pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
pass_font = Font(color="006100", bold=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for row in range(2, ws.max_row + 1):
    test_id = ws.cell(row=row, column=1).value
    if test_id in ['TS-003', 'TS-004', 'TS-005', 'TS-008']:
        ws.cell(row=row, column=10, value="PASS")
        ws.cell(row=row, column=11, value="PASS")
        ws.cell(row=row, column=12, value="Andrei")
        ws.cell(row=row, column=13, value="24.06.2026")
        ws.cell(row=row, column=14, value="Мультифакторная модель: лаги, rolling, сезонность, зарплата, выходные")
        for col in [10, 11]:
            cell = ws.cell(row=row, column=col)
            cell.fill = pass_fill
            cell.font = pass_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

wb.save('/mnt/agents/output/FitIntel_PRO_TestCases_v3.4.xlsx')
print("Excel обновлён! Мультифакторная модель в TS-003..TS-008")
