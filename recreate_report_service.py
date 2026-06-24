# recreate_report_service.py
with open('app/services/report_service.py', 'w', encoding='utf-8') as f:
    f.write('''from io import BytesIO
from datetime import datetime, date
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from sqlalchemy.orm import Session
from app.models.payment import Payment


class ReportService:
    def __init__(self, db: Session):
        self.db = db
    
    def export_payments_xlsx(
        self,
        date_from=None,
        date_to=None,
        client_id=None,
        payment_direction=None,
        payment_category=None,
        status=None,
    ):
        query = self.db.query(Payment)
        
        if date_from:
            query = query.filter(Payment.created_at >= datetime.combine(date_from, datetime.min.time()))
        if date_to:
            query = query.filter(Payment.created_at <= datetime.combine(date_to, datetime.max.time()))
        if client_id:
            query = query.filter(Payment.client_id == client_id)
        if payment_direction:
            query = query.filter(Payment.payment_direction == payment_direction)
        if payment_category:
            query = query.filter(Payment.payment_category == payment_category)
        if status:
            query = query.filter(Payment.status == status)
        
        payments = query.order_by(Payment.created_at.desc()).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Payments"
        
        headers = [
            "No", "Payment ID", "Date", "Time", "Direction", "Category",
            "Client Name", "Phone", "Email", "Amount", "Currency",
            "Method", "Type", "Bank/System", "Status",
            "Purpose", "Receipt No", "Created By"
        ]
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        for row_idx, payment in enumerate(payments, 2):
            client_name = ""
            client_phone = ""
            client_email = ""
            if payment.client:
                parts = [p for p in [payment.client.last_name, payment.client.first_name, payment.client.middle_name] if p]
                client_name = " ".join(parts)
                client_phone = payment.client.phone or ""
                client_email = payment.client.email or ""
            
            payment_type = "Cash"
            if payment.payment_method in ["CARD", "SBP", "ONLINE", "TRANSFER"]:
                payment_type = "Non-cash"
            
            bank_name = payment.payment_system or ""
            if payment.payment_method == "CASH":
                bank_name = "Cash"
            elif payment.payment_method == "SBP":
                bank_name = "SBP"
            
            receipt_num = ""
            if payment.receipt:
                receipt_num = payment.receipt.receipt_number or ""
            
            ws.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border
            ws.cell(row=row_idx, column=2, value=str(payment.id)).border = thin_border
            ws.cell(row=row_idx, column=3, value=payment.created_at.strftime("%d.%m.%Y") if payment.created_at else "").border = thin_border
            ws.cell(row=row_idx, column=4, value=payment.created_at.strftime("%H:%M:%S") if payment.created_at else "").border = thin_border
            ws.cell(row=row_idx, column=5, value=payment.payment_direction or "INCOME").border = thin_border
            ws.cell(row=row_idx, column=6, value=payment.payment_category or "SUBSCRIPTION").border = thin_border
            ws.cell(row=row_idx, column=7, value=client_name).border = thin_border
            ws.cell(row=row_idx, column=8, value=client_phone).border = thin_border
            ws.cell(row=row_idx, column=9, value=client_email).border = thin_border
            ws.cell(row=row_idx, column=10, value=float(payment.amount)).border = thin_border
            ws.cell(row=row_idx, column=11, value=payment.currency or "RUB").border = thin_border
            ws.cell(row=row_idx, column=12, value=payment.payment_method or "").border = thin_border
            ws.cell(row=row_idx, column=13, value=payment_type).border = thin_border
            ws.cell(row=row_idx, column=14, value=bank_name).border = thin_border
            ws.cell(row=row_idx, column=15, value=payment.status or "").border = thin_border
            ws.cell(row=row_idx, column=16, value=payment.notes or "").border = thin_border
            ws.cell(row=row_idx, column=17, value=receipt_num).border = thin_border
            ws.cell(row=row_idx, column=18, value=str(payment.created_by_user_id) if payment.created_by_user_id else "").border = thin_border
        
        for col_idx in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col_idx)
            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        ws.freeze_panes = 'A2'
        
        total_row = len(payments) + 2
        ws.cell(row=total_row, column=1, value="TOTAL:").font = Font(bold=True)
        ws.cell(row=total_row, column=10, value=sum(float(p.amount) for p in payments)).font = Font(bold=True)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def export_payments_csv(
        self,
        date_from=None,
        date_to=None,
        client_id=None,
        payment_direction=None,
        payment_category=None,
        status=None,
    ):
        import csv
        from io import StringIO
        
        query = self.db.query(Payment)
        
        if date_from:
            query = query.filter(Payment.created_at >= datetime.combine(date_from, datetime.min.time()))
        if date_to:
            query = query.filter(Payment.created_at <= datetime.combine(date_to, datetime.max.time()))
        if client_id:
            query = query.filter(Payment.client_id == client_id)
        if payment_direction:
            query = query.filter(Payment.payment_direction == payment_direction)
        if payment_category:
            query = query.filter(Payment.payment_category == payment_category)
        if status:
            query = query.filter(Payment.status == status)
        
        payments = query.order_by(Payment.created_at.desc()).all()
        
        output = StringIO()
        writer = csv.writer(output, delimiter=';', lineterminator='\\n')
        
        headers = [
            "No", "Payment ID", "Date", "Time", "Direction", "Category",
            "Client Name", "Phone", "Email", "Amount", "Currency",
            "Method", "Type", "Bank/System", "Status",
            "Purpose", "Receipt No", "Created By"
        ]
        writer.writerow(headers)
        
        for idx, payment in enumerate(payments, 1):
            client_name = ""
            client_phone = ""
            client_email = ""
            if payment.client:
                parts = [p for p in [payment.client.last_name, payment.client.first_name, payment.client.middle_name] if p]
                client_name = " ".join(parts)
                client_phone = payment.client.phone or ""
                client_email = payment.client.email or ""
            
            payment_type = "Cash"
            if payment.payment_method in ["CARD", "SBP", "ONLINE", "TRANSFER"]:
                payment_type = "Non-cash"
            
            bank_name = payment.payment_system or ""
            if payment.payment_method == "CASH":
                bank_name = "Cash"
            elif payment.payment_method == "SBP":
                bank_name = "SBP"
            
            receipt_num = ""
            if payment.receipt:
                receipt_num = payment.receipt.receipt_number or ""
            
            writer.writerow([
                idx,
                str(payment.id),
                payment.created_at.strftime("%d.%m.%Y") if payment.created_at else "",
                payment.created_at.strftime("%H:%M:%S") if payment.created_at else "",
                payment.payment_direction or "INCOME",
                payment.payment_category or "SUBSCRIPTION",
                client_name,
                client_phone,
                client_email,
                float(payment.amount),
                payment.currency or "RUB",
                payment.payment_method or "",
                payment_type,
                bank_name,
                payment.status or "",
                payment.notes or "",
                receipt_num,
                str(payment.created_by_user_id) if payment.created_by_user_id else "",
            ])
        
        writer.writerow([
            "TOTAL:", "", "", "", "", "", "", "", "",
            sum(float(p.amount) for p in payments),
            "", "", "", "", "", "", "", ""
        ])
        
        bytes_output = BytesIO()
        bytes_output.write(output.getvalue().encode('utf-8-sig'))
        bytes_output.seek(0)
        return bytes_output
    
    def export_payments_pdf(
        self,
        date_from=None,
        date_to=None,
        client_id=None,
        payment_direction=None,
        payment_category=None,
        status=None,
    ):
        from fpdf import FPDF
        
        query = self.db.query(Payment)
        
        if date_from:
            query = query.filter(Payment.created_at >= datetime.combine(date_from, datetime.min.time()))
        if date_to:
            query = query.filter(Payment.created_at <= datetime.combine(date_to, datetime.max.time()))
        if client_id:
            query = query.filter(Payment.client_id == client_id)
        if payment_direction:
            query = query.filter(Payment.payment_direction == payment_direction)
        if payment_category:
            query = query.filter(Payment.payment_category == payment_category)
        if status:
            query = query.filter(Payment.status == status)
        
        payments = query.order_by(Payment.created_at.desc()).all()
        
        pdf = FPDF(orientation='L', unit='mm', format='A4')
        pdf.add_page()
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'Payments Report', ln=True, align='C')
        pdf.set_font('Arial', '', 10)
        pdf.cell(0, 6, f'Period: {date_from} - {date_to}', ln=True, align='C')
        pdf.ln(5)
        
        headers = ['No', 'Date', 'Direction', 'Category', 'Client', 'Amount', 'Status', 'Purpose']
        col_widths = [10, 25, 25, 30, 50, 25, 25, 60]
        
        pdf.set_font('Arial', 'B', 8)
        pdf.set_fill_color(68, 114, 196)
        pdf.set_text_color(255, 255, 255)
        for i, h in enumerate(headers):
            pdf.cell(col_widths[i], 8, h, border=1, fill=True, align='C')
        pdf.ln()
        
        pdf.set_font('Arial', '', 8)
        pdf.set_text_color(0, 0, 0)
        total = 0
        for idx, payment in enumerate(payments, 1):
            client_name = ''
            if payment.client:
                parts = [p for p in [payment.client.last_name, payment.client.first_name] if p]
                client_name = ' '.join(parts)
            
            row = [
                str(idx),
                payment.created_at.strftime('%d.%m.%Y') if payment.created_at else '',
                payment.payment_direction or 'INCOME',
                payment.payment_category or 'SUBSCRIPTION',
                client_name,
                f'{float(payment.amount):.2f}',
                payment.status or '',
                (payment.notes or '')[:30]
            ]
            
            for i, cell in enumerate(row):
                pdf.cell(col_widths[i], 6, str(cell), border=1, align='C' if i in [0, 5] else 'L')
            pdf.ln()
            total += float(payment.amount)
        
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(sum(col_widths) - col_widths[5], 8, 'TOTAL:', border=1, align='R')
        pdf.cell(col_widths[5], 8, f'{total:.2f}', border=1, align='C')
        pdf.ln()
        
        output = BytesIO()
        pdf.output(output)
        output.seek(0)
        return output
    
    def export_payments_1c(self, date_from=None, date_to=None):
        query = self.db.query(Payment)
        
        if date_from:
            query = query.filter(Payment.created_at >= datetime.combine(date_from, datetime.min.time()))
        if date_to:
            query = query.filter(Payment.created_at <= datetime.combine(date_to, datetime.max.time()))
        
        payments = query.order_by(Payment.created_at.desc()).all()
        
        lines = []
        lines.append('1CCLIENTCOMMA')
        lines.append('###')
        lines.append('Date;Time;DocNo;DocType;Amount;Currency;Counterparty;INN;KPP;Purpose;Status')
        
        for payment in payments:
            client_name = ''
            if payment.client:
                parts = [p for p in [payment.client.last_name, payment.client.first_name] if p]
                client_name = ' '.join(parts)
            
            lines.append(
                f'{payment.created_at.strftime("%d.%m.%Y") if payment.created_at else ""};'
                f'{payment.created_at.strftime("%H:%M:%S") if payment.created_at else ""};'
                f'{str(payment.id)};'
                f'Payment;'
                f'{float(payment.amount):.2f};'
                f'{payment.currency or "RUB"};'
                f'{client_name};'
                f';;'
                f'{(payment.notes or "")[:50]};'
                f'{payment.status or ""}'
            )
        
        lines.append(f'TOTAL;{sum(float(p.amount) for p in payments):.2f}')
        
        output = BytesIO()
        output.write('\\n'.join(lines).encode('utf-8-sig'))
        output.seek(0)
        return output
    
    def export_visits_xlsx(self, date_from=None, date_to=None, client_id=None):
        from app.models.visit import Visit
        
        query = self.db.query(Visit)
        
        if date_from:
            query = query.filter(Visit.entry_time >= datetime.combine(date_from, datetime.min.time()))
        if date_to:
            query = query.filter(Visit.entry_time <= datetime.combine(date_to, datetime.max.time()))
        if client_id:
            query = query.filter(Visit.client_id == client_id)
        
        visits = query.order_by(Visit.entry_time.desc()).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Visits'
        
        headers = ['No', 'Visit ID', 'Entry Date', 'Entry Time', 'Exit Date', 'Exit Time',
                   'Duration (min)', 'Client Name', 'Phone', 'Subscription', 'Status', 'Zone', 'Access Method']
        
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        for row_idx, visit in enumerate(visits, 2):
            client_name = ''
            client_phone = ''
            if visit.client:
                parts = [p for p in [visit.client.last_name, visit.client.first_name] if p]
                client_name = ' '.join(parts)
                client_phone = visit.client.phone or ''
            
            ws.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border
            ws.cell(row=row_idx, column=2, value=str(visit.id)).border = thin_border
            ws.cell(row=row_idx, column=3, value=visit.entry_time.strftime('%d.%m.%Y') if visit.entry_time else '').border = thin_border
            ws.cell(row=row_idx, column=4, value=visit.entry_time.strftime('%H:%M:%S') if visit.entry_time else '').border = thin_border
            ws.cell(row=row_idx, column=5, value=visit.exit_time.strftime('%d.%m.%Y') if visit.exit_time else '').border = thin_border
            ws.cell(row=row_idx, column=6, value=visit.exit_time.strftime('%H:%M:%S') if visit.exit_time else '').border = thin_border
            ws.cell(row=row_idx, column=7, value=visit.duration_minutes or 0).border = thin_border
            ws.cell(row=row_idx, column=8, value=client_name).border = thin_border
            ws.cell(row=row_idx, column=9, value=client_phone).border = thin_border
            ws.cell(row=row_idx, column=10, value='').border = thin_border
            ws.cell(row=row_idx, column=11, value=visit.status or '').border = thin_border
            ws.cell(row=row_idx, column=12, value=visit.zone or '').border = thin_border
            ws.cell(row=row_idx, column=13, value=visit.access_method or '').border = thin_border
        
        for col_idx in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col_idx)
            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        ws.freeze_panes = 'A2'
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def export_visits_csv(self, date_from=None, date_to=None, client_id=None):
        import csv
        from io import StringIO
        from app.models.visit import Visit
        
        query = self.db.query(Visit)
        
        if date_from:
            query = query.filter(Visit.entry_time >= datetime.combine(date_from, datetime.min.time()))
        if date_to:
            query = query.filter(Visit.entry_time <= datetime.combine(date_to, datetime.max.time()))
        if client_id:
            query = query.filter(Visit.client_id == client_id)
        
        visits = query.order_by(Visit.entry_time.desc()).all()
        
        output = StringIO()
        writer = csv.writer(output, delimiter=';', lineterminator='\\n')
        
        headers = ['No', 'Visit ID', 'Entry Date', 'Entry Time', 'Exit Date', 'Exit Time',
                   'Duration (min)', 'Client Name', 'Phone', 'Subscription', 'Status', 'Zone', 'Access Method']
        writer.writerow(headers)
        
        for idx, visit in enumerate(visits, 1):
            client_name = ''
            client_phone = ''
            if visit.client:
                parts = [p for p in [visit.client.last_name, visit.client.first_name] if p]
                client_name = ' '.join(parts)
                client_phone = visit.client.phone or ''
            
            writer.writerow([
                idx,
                str(visit.id),
                visit.entry_time.strftime('%d.%m.%Y') if visit.entry_time else '',
                visit.entry_time.strftime('%H:%M:%S') if visit.entry_time else '',
                visit.exit_time.strftime('%d.%m.%Y') if visit.exit_time else '',
                visit.exit_time.strftime('%H:%M:%S') if visit.exit_time else '',
                visit.duration_minutes or 0,
                client_name,
                client_phone,
                '',
                visit.status or '',
                visit.zone or '',
                visit.access_method or ''
            ])
        
        bytes_output = BytesIO()
        bytes_output.write(output.getvalue().encode('utf-8-sig'))
        bytes_output.seek(0)
        return bytes_output
''')

print("report_service.py пересоздан!")
