# app/services/export_service.py
import csv
import json
import os
from datetime import datetime, date
from decimal import Decimal
from typing import List, Dict, Optional
from uuid import UUID

from sqlalchemy.orm import Session
from sqlalchemy import text

from app.models.export import ExportJob


class ExportService:
    """Сервис экспорта данных в разные форматы"""

    def __init__(self, db: Session, output_dir: str = "exports"):
        self.db = db
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def create_job(self, club_id: int, user_id: UUID, entity: str, format: str, filters: dict) -> ExportJob:
        """Создать задачу экспорта"""
        job = ExportJob(
            club_id=club_id,
            user_id=user_id,
            entity=entity,
            format=format,
            status='pending',
            filters=filters
        )
        self.db.add(job)
        self.db.commit()
        self.db.refresh(job)
        return job

    def export(self, job_id: UUID) -> str:
        """Выполнить экспорт"""
        job = self.db.query(ExportJob).filter(ExportJob.id == job_id).first()
        if not job:
            raise ValueError("Job not found")

        job.status = 'processing'
        self.db.commit()

        try:
            entity = job.entity
            format_type = job.format
            filters = job.filters or {}

            # Получаем данные
            data = self._fetch_data(entity, filters)

            # Генерируем файл
            filename = f"export_{entity}_{job_id}.{format_type}"
            filepath = os.path.join(self.output_dir, filename)

            if format_type == 'xlsx':
                self._export_xlsx(data, filepath)
            elif format_type == 'csv':
                self._export_csv(data, filepath)
            elif format_type == 'json':
                self._export_json(data, filepath)
            elif format_type == 'xml':
                self._export_xml(data, filepath)
            else:
                raise ValueError(f"Unsupported format: {format_type}")

            # Обновляем job
            job.status = 'completed'
            job.file_path = filepath
            job.row_count = len(data)
            job.file_size = os.path.getsize(filepath)
            job.completed_at = datetime.now()
            self.db.commit()

            return filepath

        except Exception as e:
            job.status = 'failed'
            job.error_message = str(e)
            self.db.commit()
            raise

    def _fetch_data(self, entity: str, filters: dict) -> List[Dict]:
        """Получить данные из БД"""
        club_id = filters.get('club_id')
        date_from = filters.get('date_from')
        date_to = filters.get('date_to')

        if entity == 'clients':
            query = text("""
                SELECT id, first_name, last_name, phone, email, birth_date, status, created_at
                FROM clients
                ORDER BY created_at DESC
                LIMIT 1000
            """)
            result = self.db.execute(query).mappings().all()

        elif entity == 'payments':
            query = text("""
                SELECT p.id, p.amount, p.status, p.payment_method, p.created_at,
                       c.first_name, c.last_name
                FROM payments p
                JOIN clients c ON c.id = p.client_id
                WHERE p.club_id = :club_id
                AND p.created_at BETWEEN :date_from AND :date_to
                ORDER BY p.created_at DESC
            """)
            result = self.db.execute(query, {
                "club_id": club_id,
                "date_from": date_from or '1900-01-01',
                "date_to": date_to or '2099-12-31'
            }).mappings().all()

        elif entity == 'visits':
            query = text("""
                SELECT v.id, v.visit_date, v.check_in, v.check_out, v.status,
                       c.first_name, c.last_name
                FROM visits v
                JOIN clients c ON c.id = v.client_id
                WHERE v.club_id = :club_id
                AND v.visit_date BETWEEN :date_from AND :date_to
                ORDER BY v.visit_date DESC
            """)
            result = self.db.execute(query, {
                "club_id": club_id,
                "date_from": date_from or '1900-01-01',
                "date_to": date_to or '2099-12-31'
            }).mappings().all()

        elif entity == 'sales':
            query = text("""
                SELECT s.id, s.product_name, s.amount, s.commission_amount, s.sale_type, s.created_at,
                       c.first_name, c.last_name
                FROM trainer_sales s
                JOIN clients c ON c.id = s.client_id
                WHERE s.created_at BETWEEN :date_from AND :date_to
                ORDER BY s.created_at DESC
            """)
            result = self.db.execute(query, {
                "date_from": date_from or '1900-01-01',
                "date_to": date_to or '2099-12-31'
            }).mappings().all()

        else:
            raise ValueError(f"Unknown entity: {entity}")

        return [dict(row) for row in result]

    def _export_csv(self, data: List[Dict], filepath: str):
        """Экспорт в CSV"""
        if not data:
            with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
                f.write('')
            return

        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=data[0].keys())
            writer.writeheader()
            for row in data:
                writer.writerow(self._serialize_row(row))

    def _export_json(self, data: List[Dict], filepath: str):
        """Экспорт в JSON"""
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump([self._serialize_row(row) for row in data], f, ensure_ascii=False, indent=2, default=str)

    def _export_xlsx(self, data: List[Dict], filepath: str):
        """Экспорт в Excel (XLSX)"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data"

            if data:
                # Заголовки
                headers = list(data[0].keys())
                ws.append(headers)
                
                # Стили заголовков
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

                # Данные
                for row in data:
                    ws.append(list(self._serialize_row(row).values()))

                # Автоширина
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(filepath)
        except ImportError:
            # Fallback: CSV с расширением .xlsx
            self._export_csv(data, filepath)

    def _export_xml(self, data: List[Dict], filepath: str):
        """Экспорт в XML (формат 1C)"""
        from xml.etree.ElementTree import Element, SubElement, tostring
        from xml.dom.minidom import parseString

        root = Element("ФитнесКлуб")
        root.set("Версия", "1.0")
        root.set("ДатаФормирования", datetime.now().isoformat())

        for row in data:
            item = SubElement(root, "Запись")
            for key, value in row.items():
                field = SubElement(item, self._xml_safe_name(key))
                field.text = str(value) if value is not None else ""

        xml_str = tostring(root, encoding='utf-8')
        pretty_xml = parseString(xml_str).toprettyxml(indent="  ", encoding='utf-8')
        
        with open(filepath, 'wb') as f:
            f.write(pretty_xml)

    def _serialize_row(self, row: Dict) -> Dict:
        """Сериализация строки (Decimal, UUID, datetime)"""
        result = {}
        for key, value in row.items():
            if isinstance(value, Decimal):
                result[key] = float(value)
            elif isinstance(value, UUID):
                result[key] = str(value)
            elif isinstance(value, (datetime, date)):
                result[key] = value.isoformat()
            else:
                result[key] = value
        return result

    def _xml_safe_name(self, name: str) -> str:
        """Безопасное имя для XML"""
        return name.replace(' ', '_').replace('-', '_').replace('.', '_')
